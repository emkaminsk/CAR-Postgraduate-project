# -*- coding: utf-8 -*-
"""
Created on Sun Mar 12 17:03:37 2017

@author: Marcin

program do uruchomienia wymaga podfolderow project_path:
    'CMT' - pojedyncze CMT do analizy
    'CMT_part' - wybrane CMT do analizy (opcjonalnie)
    'CMT_all' - komplet CMT do analizy (opcjonalnie)
    'py' - kody do uruchomienia
    'logs' - logi
    
    Wynikiem działania jest plik car.graphml do analizy np. w Gephi
"""

from openpyxl import load_workbook
from pathlib import Path
import networkx as nx
import matplotlib.pyplot as plt
import re
import datetime

def build_car_graph(g, ark, nazwa):  
    """funkcja rozbudowuje graf g o dane ze wskazanego arkusza (zakładki) CMT 
    Indeks oznaczen w atrybucie 'sql' w wierzchołkach grafu - opis czym jest obiekt
    CAR - wynikowe tabele CAR, jest ich ok. 30
    CAR_part - częsci wynikowych tabel CAR, zakładki w CMT, w tym zakładki techniczne SUB czy archiwa
    Source - tabela bedaca źrodlem zasilania CAR
    Field - pole w tabeli
    
    Indeks oznaczen w atrybucie 'sql' w krawedziach grafu - opis działania obiektu na obiekt
    Merge - polaczenie CAR_part do CAR
    From - polaczenie miedzy tabelami zrodlowymi a tabela wynikowa CAR_part
    In - polaczenie mowiace, ze pole znajduje sie w tabeli
    Transform - polaczenie miedzy polami zrodlowymi a wynikowymi
    SelCond, OutSelCond, Having, Qualify, Groupby - relacja od pola do zakładki w CMT     """              
    output_sheet = '${CAR_TEMP}.'+ nazwa if nazwa.startswith('SUB_') else '${CAR}.' + nazwa 
    output_table = cl(ark,3,10)
    output_table_raw = output_table[7:]
    log(fhlog, '\nArkusz: ' + output_sheet + ' Tabela:' + output_table)
    
    #odczyt CMT i edycja grafu
    
    #tabele From
    tables_in_keys = [cl(ark,x,2) for x in range(3,12) if cl(ark,x,col_table) != '']    
    tables_in_vals = [cmt_field_table_correct(cl(ark,x,col_table),output_table_raw) for x in range(3,12) if cl(ark,x,col_table) != '']    
    tables_in = dict(zip(tables_in_keys,tables_in_vals))
    log(fhlog, 'Tabele:', tables_in)
    
    g.add_node(output_table, sql='CAR')
    g.add_node(output_sheet, sql='CAR_part')
    g.add_edge(output_sheet, output_table, Label='Merge', sql='Merge')
    i = 1
    for klucz,tabela in tables_in.items():        
        if not tabela.startswith('${CAR'):
            g.add_node(tabela, sql='Source')
        elif tabela.startswith('${CAR_TEMP}'):
            g.add_node(tabela, sql='CAR_part') if not g.has_node(tabela) else 1==1
        elif tabela.startswith('${CAR}') and not tabela.startswith(output_table):
            g.add_node(tabela, sql='CAR') if not g.has_node(tabela) else 1==1
        else: 
            log(fhlog,'Bez node: ' + tabela)
        g.add_edge(tabela, output_sheet, Label='From', sql='From', From_id = i)
        i = i + 1
    
    #SEL_COND 13-21 - warunki ograniczające przed złączeniem
    sel_cond = [cl(ark,x,col_table) for x in range(13,21) if cl(ark,x,col_table) != '']    
    if len(sel_cond) != 0:
        for elem in sel_cond:   
            cmt_extend_graph(elem, g, output_sheet, tables_in, 'Sel_Cond', fhlog)
            
    #OUT_SEL_COND 22 - warunki ograniczające po złączeniu
    out_sel_cond = cl(ark,22,col_table)     
    if out_sel_cond != '':
        cmt_extend_graph(out_sel_cond, g, output_sheet, tables_in, 'Out_Sel_Cond', fhlog)
                
    #GROUPBY 23
    groupby = cl(ark,23,col_table)
    if groupby != '':    
        cmt_extend_graph(groupby, g, output_sheet, tables_in, 'Group_By', fhlog)
            
    #HAVING 24
    having = cl(ark,24,col_table)
    if having != '':
        cmt_extend_graph(having, g, output_sheet, tables_in, 'Having', fhlog)
    
    #QUALIFY 25
    qualify = cl(ark,25,col_table)
    if qualify != '':
        cmt_extend_graph(qualify, g, output_sheet, tables_in, 'Qualify', fhlog)
        
    #pola    
    a = 31
    target_fields = {}
    target_fields_output = {}
    while cl(ark,a,12) != '':
        if  cl(ark,a,12) != 'N/D':
            target_fields[a-30] = output_sheet + '.' + cl(ark,a,12)
            target_fields_output[a-30] = output_table + '.' + cl(ark,a,12)
        a = a + 1        
    log(fhlog, 'Target fields:', target_fields)

    a = 31
    source_fields = {}
    while cl(ark,a,3) != '':
        if  cl(ark,a,3) != 'N/D':
            #source_fields[a-30] = [tables_in[cl(ark,a,5)], cl(ark,a,3) + '.' + cl(ark,a,4) + '.' + cl(ark,a,6)]
            source_fields[a-30] = [tables_in[cl(ark,a,5)], tables_in[cl(ark,a,5)] + '.' + cl(ark,a,6)]
        a = a + 1    
    log(fhlog, 'Source fields:', source_fields)
    
    for x in range(1, a-30):
        try:
            g.add_node(source_fields[x][1], sql='Field')
            g.add_edge(source_fields[x][1], source_fields[x][0], Label='In', sql = 'In')
            if x in target_fields:
                g.add_node(target_fields[x], sql='Field')
                g.add_edge(target_fields[x], output_sheet, Label='In',sql='In')
                g.add_node(target_fields_output[x], sql='Field')
                g.add_edge(target_fields[x], target_fields_output[x], Label='Identity', sql='Identity')
                y = x
            g.add_edge(source_fields[x][1], target_fields[y], Label='Transform', sql='Transform')
        except KeyError:
            pass
        
    return g        

def cl(ark, a, b):
    """funkcja sluzy jako skrot do łatwego odczytu konkretnej komorki z CMT"""
    try:
        return str(ark.cell(row=a, column=b).value).strip() if ark.cell(row=a, column=b).value != None else ''
    except IndexError:
        return ''

def cl_addr(ark, a, b):
    """funkcja na zapas, zwraca adres komorki"""
    return str(ark.cell(row=a, column=b).coordinate)

def cmt_field_table_correct(tabela, nazwa):
    """ funkcja do zamiany przedrostka na własciwy
    w CMT TABLE_MASTER i inne pola w tej sekcji dla zakładek CAR zamiast ${CAR} mają ${CAR_TEMP} """
    if tabela.startswith('${CAR_TEMP}') and not tabela.startswith('${CAR_TEMP}.SUB') and bool(re.search(nazwa + r'_\d+$',tabela)):
        return '${CAR}.CMT_' + nazwa + '_' + tabela.split('_')[-1]
    else:
        return tabela

def cmt_table_translate(pola, tables_in):
    """funkcja do zamiany klucza na wartosc wg słownika tables_in
    pierwszy parametr pola to lista, drugi to slownik"""
    wynik = list()
    for elem in pola:
        baza = elem.split('.')[0]
        if baza in tables_in:
            wynik.append(elem.replace(baza,tables_in[baza]))
        else:
            elem_temp = re_translate.sub('', elem)
            for k,v in tables_in.items():
                elem_temp = re.sub(k,v,elem_temp)
            wynik.append(elem_temp)                
    return wynik
        
def cmt_extract_field_list(tekst):
    """funkcja przyjmuje fragment kwerendy i wyciąga z niego pola"""
    return re_fields.findall(tekst)

def cmt_extend_graph(sql_part, g, target_table, tables_in, attr_name, fhlog):
    """standardowa funkcja uzupelniajaca graf o pola wyciągniete z fragmentu sql (sql_part)
    attr_name - nazwa atrybutu, po ktorym będzie można pole znalezc w grafie
    do zmiany -- zamiast sql musi byc wiele flag osobnych dla roznych atrybutow"""
    log(fhlog, attr_name + ': ' + sql_part)
    fields = cmt_table_translate(cmt_extract_field_list(sql_part), tables_in)
    log(fhlog, attr_name + ' Fields: ', fields)
    for field in fields:
        g.add_node(field, sql='Field')
        g.add_edge(field, target_table, Label=attr_name, sql=attr_name)

def log(fh, *tekst):    
    """funkcja do zapisywania do pliku logow z działania programu"""
    for tk in tekst:
        fh.write(str(tk))
    fh.write('\n')
                
def draw(g):
    """rysowanie   """      
    #nx.draw_circular(g)
    nx.draw_random(g)
    #spectral, shell, spring
    plt.show()   

def save_graph(g,project_path, filename):
    nx.write_graphml(g,project_path + filename + ".graphml")
        
# main 
#if __name__  == '__main__':       
def CMT_parser(progress_Bar=None, project_path = 'd:\\projekty\\zajęcia na uczelni\\!tpd\\', input_zakres ='CMT_all', filename = 'car'):  
    start_time = datetime.datetime.now()
     #common vars
    global col_table, re_groupby, re_translate, re_fields, fhlog
    #project_path = "d:\\projekty\\zajęcia na uczelni\\!tpd\\"
    #zakres = 'CMT\\'
    #zakres = 'CMT_part\\'
    zakres = input_zakres + '\\'
    col_table = 5 #głowna kolumna gdzie jest wiekszosc danych w CMT
    #skompilowne wyrazenia regularne
    re_groupby = re.compile(r'[ \(\)]')
    re_translate = re.compile(r'[\t\n\r\f\v]')
    re_fields = re.compile(r'\bTABLE\w+?\b\.\b\w+?\b')
    
    #build    
    g = nx.DiGraph()    
    p = Path(project_path + zakres)  
    x = 0
    total = float(len(list(p.glob('*.xlsx'))))
    for filename in p.glob('*.xlsx'):
        if str(filename).count('CMT_DICT_PARAM_STATIC') > 0:
            continue
        else:
            print('Loading... ' + str(filename))
            wb = load_workbook(str(filename), read_only=True, data_only=True)
            for sheet in wb:   
                if ["Title_page", "Document_history"].count(sheet.title) == 0:
                    sheet.max_row = sheet.max_column = None
                    fhlog = open(project_path+'logs\\log' + sheet.title + '.txt', mode='w')
                    g = build_car_graph(g, sheet, sheet.title)                
                    fhlog.close()                
            wb._archive.close()
        x +=1
        if not progress_Bar is None:
            progress_Bar.setValue(100*x/total)
    #draw(g)
    save_graph(g, project_path, filename)
    end_time = datetime.datetime.now()
    return 'Zakończone. Całkowity czas: ' + str(end_time - start_time)
